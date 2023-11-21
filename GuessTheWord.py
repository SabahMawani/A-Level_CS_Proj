import time
import random
import openpyxl as px
import pandas as pd
Option = pd.Series(['1','2','3','4','5','6'])
Letters = pd.Series(['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z'])
def Country():
    index = random.randint(1,20)
    index = str(index)
    book = px.load_workbook('HangMan.xlsx')
    sheet = book['Country']
    cell='A'+index
    word = sheet[cell].value
    return (word)
def Bands():
    index = random.randint(1,20)
    index = str(index)
    book = px.load_workbook('HangMan.xlsx')
    sheet = book['Bands']
    cell='A'+index
    word = sheet[cell].value
    return (word)
def Movies():
    index = random.randint(1,20)
    index = str(index)
    book = px.load_workbook('HangMan.xlsx')
    sheet = book['Movies']
    cell='A'+index
    word = sheet[cell].value
    return (word)
def Books():
    index = random.randint(1,20)
    index = str(index)
    book = px.load_workbook('HangMan.xlsx')
    sheet = book['Books']
    cell='A'+index
    word = sheet[cell].value
    return (word)
def Animals():
    index = random.randint(1,20)
    index = str(index)
    book = px.load_workbook('HangMan.xlsx')
    sheet = book['Animals']
    cell='A'+index
    word = sheet[cell].value
    return (word)
def Colors():
    index = random.randint(1,20)
    index = str(index)
    book = px.load_workbook('HangMan.xlsx')
    sheet = book['Colors']
    cell='A'+index
    word = sheet[cell].value
    return (word)
def choice():
    print("Which topic would you like to choose?")
    print("1. Movies")
    print("2. Books")
    print("3. Animals")
    print("4. Colors")
    print("5. Countries")
    print("6. Rock Bands")
    time.sleep(0.75)
    print("Enter the number corresponding to your choice.")
    time.sleep(0.75)
    Choice = input("Enter your choice here: ")
    while Choice not in Option.values:
        print("Please enter a valid choice.")
        Choice = input("Enter your choice here: ")
    if Choice == '1':
        word = Movies()
    elif Choice == '2':
        word = Books()
    elif Choice == '3':
        word = Animals()
    elif Choice == '4':
        word = Colors()
    elif Choice == '5':
        word = Country()
    elif Choice == '6':
        word = Bands()
    return (word)
def Continue():
    option = ""
    print ("Would you like to play again? (y/n)")
    option = input("Enter your choice here: ")
    option = option.upper()
    Flag3 = False
    while Flag3 == False:
        if option == "N":
            Flag = False
            Flag3 = True
        elif option == "Y":
            Flag = True
            Flag3 = True
        else:
            print("Invalid input, please enter again.")
            option = input("Enter your choice here: ")
            option = option.upper()
    
    return Flag
def main():
    print(" ")
    time.sleep(1)
    word = choice()
    word = word.upper()
    print("Start guessing...")
    time.sleep(0.5)
    guesses = ' :'
    turns = 10
    while turns >0:
        failed = 0
        for char in word:
            if char in guesses:
                print (char, end=" ")
            else:
                print("_", end=" ")
                failed = failed + 1
        if failed == 0:
            print('\n')
            print("Congratulations")
            print("You Won,",name)
            break
        print ('\n')
        print ("You have ",turns,"more guesses.")
        print ("Alphabets that have already been entered",guesses)
        Flag1 = False
        Flag2 = False
        while Flag1 == False or Flag2 == False:
            guess = input("Guess an alphabet: ")
            guess = guess.upper()
            if len(guess) != 1:
                print("Please enter one character.")
            else:
                Flag2 = True
            if guess not in Letters.values:
                print("Please enter an alphabet.")
            elif guess in guesses:
                print("This alphabet has already been entered before.")
            else:
                Flag1 = True
        guesses +=guess
        if guess not in word:
            turns = turns-1
            print ("Wrong")
            if turns == 0:
                print("You Lose!")
                print("Better luck next time,",name)
                print("The word was, ",word)
Flag = True
print("What is your name?")
name = input("Enter your name here: ")
print("Hello, ",name,", Time to play hangman.")
while Flag == True:
    main()
    time.sleep(0.5)
    Flag = Continue()
    
